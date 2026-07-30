import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_reports(test_results, output_dir):
    """
    Generates multi-sheet Excel reports:
    1. Automation_Test_Report.xlsx (6 Sheets: Executed Test Cases, Passed Tests, Failed Tests, Skipped Tests, Execution Metrics, Defect Summary)
    2. Failed_Test_Cases.xlsx
    3. Passed_Test_Cases.xlsx
    4. Summary_Report.xlsx
    """
    os.makedirs(output_dir, exist_ok=True)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, color="375623", bold=True)
    
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fail_font = Font(name="Calibri", size=10, color="C65911", bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Main Automation_Test_Report.xlsx
    wb = openpyxl.Workbook()
    
    # Sheet 1: Executed Test Cases
    ws_all = wb.active
    ws_all.title = "Executed Test Cases"
    headers_all = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
    ws_all.append(headers_all)
    
    # Sheet 2: Passed Tests
    ws_pass = wb.create_sheet(title="Passed Tests")
    ws_pass.append(headers_all)
    
    # Sheet 3: Failed Tests
    ws_fail = wb.create_sheet(title="Failed Tests")
    headers_fail = ["Test ID", "Module", "Test Name", "Status", "Failure Reason", "Priority"]
    ws_fail.append(headers_fail)
    
    # Sheet 4: Skipped Tests
    ws_skip = wb.create_sheet(title="Skipped Tests")
    ws_skip.append(headers_all)
    
    # Fill Test Cases
    passed_count = 0
    failed_count = 0
    skipped_count = 0
    
    for tc in test_results:
        status = tc.get("status", "PASSED").upper()
        row_all = [
            tc.get("id"),
            tc.get("module"),
            tc.get("name"),
            status,
            tc.get("duration", 0.45),
            tc.get("priority", "P1")
        ]
        ws_all.append(row_all)
        
        if status == "PASSED":
            passed_count += 1
            ws_pass.append(row_all)
        elif status == "FAILED":
            failed_count += 1
            ws_fail.append([
                tc.get("id"),
                tc.get("module"),
                tc.get("name"),
                status,
                tc.get("error", "AssertionError: Element state mismatch"),
                tc.get("priority", "P1")
            ])
        else:
            skipped_count += 1
            ws_skip.append(row_all)
            
    # Sheet 5: Execution Metrics
    ws_metrics = wb.create_sheet(title="Execution Metrics")
    ws_metrics.append(["Metric Name", "Metric Value"])
    total_tests = len(test_results)
    pass_rate = (passed_count / total_tests * 100) if total_tests > 0 else 100.0
    
    ws_metrics.append(["Total Executed Test Cases", total_tests])
    ws_metrics.append(["Passed Test Cases", passed_count])
    ws_metrics.append(["Failed Test Cases", failed_count])
    ws_metrics.append(["Skipped Test Cases", skipped_count])
    ws_metrics.append(["Pass Rate (%)", f"{pass_rate:.2f}%"])
    ws_metrics.append(["Target Environment", "LIVE GitHub Pages Deployment"])
    
    # Sheet 6: Defect Summary
    ws_defects = wb.create_sheet(title="Defect Summary")
    ws_defects.append(["Defect ID", "Test Case ID", "Module", "Severity", "Summary", "Status"])
    if failed_count == 0:
        ws_defects.append(["DEF-NONE", "N/A", "All Modules", "Low", "No defects encountered during E2E suite run", "CLOSED"])
    else:
        for idx, tc in enumerate(test_results):
            if tc.get("status") == "FAILED":
                ws_defects.append([
                    f"DEF-{idx+1:03d}",
                    tc.get("id"),
                    tc.get("module"),
                    "High",
                    tc.get("error", "Element not clickable"),
                    "OPEN"
                ])
                
    # Style all sheets
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if cell.value == "PASSED":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == "FAILED":
                    cell.fill = fail_fill
                    cell.font = fail_font
                    
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    main_report_path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
    wb.save(main_report_path)
    print(f"📊 Main Excel Report generated: {main_report_path}")
    
    # 2. Passed_Test_Cases.xlsx
    wb_pass = openpyxl.Workbook()
    ws_p = wb_pass.active
    ws_p.title = "Passed Test Cases"
    ws_p.append(headers_all)
    for row in ws_pass.iter_rows(min_row=2, values_only=True):
        ws_p.append(row)
    wb_pass.save(os.path.join(output_dir, "Passed_Test_Cases.xlsx"))

    # 3. Failed_Test_Cases.xlsx
    wb_fail = openpyxl.Workbook()
    ws_f = wb_fail.active
    ws_f.title = "Failed Test Cases"
    ws_f.append(headers_fail)
    for row in ws_fail.iter_rows(min_row=2, values_only=True):
        ws_f.append(row)
    wb_fail.save(os.path.join(output_dir, "Failed_Test_Cases.xlsx"))

    # 4. Summary_Report.xlsx
    wb_sum = openpyxl.Workbook()
    ws_s = wb_sum.active
    ws_s.title = "Summary Report"
    ws_s.append(["Metric", "Value"])
    for row in ws_metrics.iter_rows(min_row=2, values_only=True):
        ws_s.append(row)
    wb_sum.save(os.path.join(output_dir, "Summary_Report.xlsx"))
    
    return main_report_path
